import math
import os
import random

import openpyxl
import xlsxwriter


# initialize excel file
wbWrite = xlsxwriter.Workbook("Distances.xlsx")
sheetWrite = wbWrite.add_worksheet()

wbRead = openpyxl.load_workbook(filename='Distances.xlsx', data_only=True)
sheetRead = wbRead.get_sheet_by_name("Sheet1")

columns = ["X1", "Y1", "X2", "Y2"]          # list for column names

pointCols = 4                               # number of columns that the coordinate data points' information exists in
pointCount = 20                             # variable for number of data points


def writeCols(cols):
   for j in range(len(cols)):
       sheetWrite.write(0, j, cols[j])


# populates excel file with random values for data points
def randomPoints():
   for i in range(1, pointCount):
       for j in range(len(columns)):
           sheetWrite.write(i, j, random.randint(1, 100))


# calculate and write Euclidean distances for coordinates
def eucDist():
   # list for the calculated euc distances
   eucDistances = []

   for i in range(1, pointCount):
       coords = []
       for j in range(pointCols):
           cell = sheetRead.cell(i+1, j+1)
           coords.append(cell.value)
       print(str(coords))
       x1 = coords[0]
       y1 = coords[1]
       x2 = coords[2]
       y2 = coords[3]
       euc = math.sqrt(math.pow((x2 - x1), 2) + math.pow((y2 - y1), 2))
       eucDistances.append(euc)

   columns.append("Euclidean Distance")
   for i in range(1, pointCount):
       sheetWrite.write(i, columns.index("Euclidean Distance"), eucDistances[i-1])



writeCols(columns)
randomPoints()
eucDist()
writeCols(columns)

wbWrite.close()

